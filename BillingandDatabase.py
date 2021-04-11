
from openpyxl import Workbook
from openpyxl import load_workbook
class Customer:
    def __init__(self, name, address, tel) :
        self.name = name
        self.address = address
        self.tel = tel

print("Welcome to Richard Hardware Store")

Question =input ("Do you want to create an invoice ? \n (Y | N) ")
if Question in ("Yy"):
    name = input("Please enter customer's name ")
    address = input("Please enter customer's address ")
    tel = input("Please enter customer's tel ")
    Customer1 = Customer(name,address,tel)
    item_qty = int(input("How many items are you going to add?"))
    i=1
    totalprice = 0
    x=0

    for i in range(item_qty):
        x =x + 1
        description = input(str(x) +". Description of item = " )
        qty = int(input(str(x) + ". Quantity = " ))
        UnitPrice = int(input(str(x) + ". Unit Price " ))
        totalprice = totalprice + (qty * UnitPrice)
        from datetime import date
        today = date.today()

        # Export to excel
        wb = Workbook()
        wb = load_workbook(filename='Invoice_DataBase.xlsx')
        ws = wb.active
        ws.append([name,address,tel,today,description,qty,UnitPrice,totalprice])


        wb.save('Invoice_DataBase.xlsx')

else:
    print("Good bye")
